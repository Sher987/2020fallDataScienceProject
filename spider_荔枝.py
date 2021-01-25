# coding=utf-8
import sys
from bs4 import BeautifulSoup
import re
import urllib
import urllib.request
import xlwt
import socket
import openpyxl

def main():
    baseurl="https://so.jstv.com/?keyword=%E7%96%AB%E6%83%85&page="   #爬取荔枝网新闻  搜索关键词：疫情
    datalist = getData(baseurl)
    # print(datalist)
    savepath=".\\.荔枝新闻.xlsx"
    saveData(datalist,savepath)

findLink=re.compile(r'<a href="(.*?)" target="_blank">')
finfTitle=re.compile(r'<title>(.*?)</title>')
findcontury=re.compile(r'<a href="/gj/">(.*?)</a>')
findcontent=re.compile(r'<div class="content">(.*?)</div>',re.S)
findtime=re.compile(r'<p class="info fL" style=" width:500px;"><span class="time">(.*?)</span>')



def getData(baseurl):
    datalist=[]
    for i in range(1180,3800):  #2019年12月8日到2020年6月20日
        html=askURL(baseurl+str(i))
        #print(html)

        # 2.对单个页面逐一进行解析
        soup = BeautifulSoup(html, "html.parser")
        #print(soup)
        for item in soup.find_all('div', class_="lzxw_per_r"):  # 查找符合要求的内容，形成列表
            item = str(item)
            #print("...")
            # print(item)
            # break

            #新闻详情页链接
            link = re.findall(findLink, item)  # 通过正则表达式来查找指定的字符串
            if(len(link)!=0):   #部分新闻没有详情链接...
                data=[]
                link=link[0]
                html=askURL(link)
                soup=BeautifulSoup(html,"html.parser")
                # print(soup)
                # return
                soupstr=str(soup)
                contury=re.findall(findcontury,soupstr)
                if(len(contury)!=0):
                    continue
                title=re.findall(finfTitle,soupstr)
                if(len(title)!=0):
                # print(soupstr)
                    time=re.findall(findtime,soupstr)
                    print(time)
                    content=str(re.findall(findcontent,soupstr)).replace(r'\u3000'," ").replace(r'\xa0',"")
                    #print(content)
                    content=re.findall('[\u4e00-\u9fa5.]*',content)
                    contentstr=""
                    for item in content:
                        contentstr+=item
                    if contentstr.__contains__("疫" or "武汉" or "湖北"):
                        #print(contentstr)
                        data.append(time)  # 发布时间
                        data.append(title[0].replace("_荔枝网新闻", " "))  # 新闻标题
                        data.append(contentstr)  #新闻内容
                        datalist.append(data)
        # print(datalist)
    return datalist

def saveData(datalist,savepath):
    print("saving……")
    wb= openpyxl.load_workbook(savepath)  # 打开现有的excel
    # book = xlwt.Workbook(encoding="utf-8", style_compression=0)  # 创建workbook对象 样式压缩
    ws = wb.create_sheet(0)  # 插在第一个位置
    # sheet = book.add_sheet('荔枝新闻内容', cell_overwrite_ok=True)  # 创建工作表  单元可覆盖
    ws['A0']="发布时间"
    ws['A1']="新闻标题"
    ws['A2']="新闻内容"

    # for i in range(0, 3):
    #     sheet.write(0, i, col[i])  # 列名
    for row in range(1,len(datalist)+1):
        data=datalist[row-1]
        for column in range (0,3):
            c=ws.cell(row=row,column=column)
            c=data[column]

    # for i in range(0, len(datalist)):
    #     data = datalist[i]
    #     for j in range(0, 3):
    #         sheet.write(i + 1, j, data[j])

    wb.save(savepath)

# 得到一个指定网页的信息
def askURL(url):
    # 实现用户代理，告诉豆瓣我们是什么类型的机器，(做到了模拟浏览器，实际上是告诉浏览器我们可以接受什么水平的信息
    head = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/87.0.4280.88 Safari/537.36"
        }
    request = urllib.request.Request(url=url, headers=head)
    html = ""
    try:
        response = urllib.request.urlopen(request,timeout=3)
        html = response.read().decode("utf-8")
        #print(html)
    except urllib.error.URLError as e:
        if (hasattr(e, "code")):
            print(e.code)
        if (hasattr(e, "reason")):
            print(e.reason)
    except socket.timeout:
        print('Time Out!')

    return html


if __name__ == '__main__':
    main()
    print("succeed!")

